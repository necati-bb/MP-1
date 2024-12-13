import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from datetime import datetime


class StudentGPAApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Student GPA and Ranking")


        self.export_dir = "exports"
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)


        self.create_widgets()

    def create_widgets(self):

        self.lbl_open = tk.Label(self.root, text="Open file:")
        self.lbl_open.grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.btn_browse = tk.Button(self.root, text="Browse", command=self.load_file)
        self.btn_browse.grid(row=0, column=1, padx=5, pady=5)


        self.lbl_id = tk.Label(self.root, text="ID:")
        self.lbl_id.grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.entry_id = tk.Entry(self.root, width=30)
        self.entry_id.grid(row=1, column=1, padx=5, pady=5)

        self.lbl_name = tk.Label(self.root, text="Name Surname:")
        self.lbl_name.grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.entry_name = tk.Entry(self.root, width=30, state="readonly")
        self.entry_name.grid(row=2, column=1, padx=5, pady=5)

        self.lbl_gpa = tk.Label(self.root, text="GPA:")
        self.lbl_gpa.grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.entry_gpa = tk.Entry(self.root, width=30, state="readonly")
        self.entry_gpa.grid(row=3, column=1, padx=5, pady=5)

        self.lbl_rank = tk.Label(self.root, text="Rank:")
        self.lbl_rank.grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.entry_rank = tk.Entry(self.root, width=30, state="readonly")
        self.entry_rank.grid(row=4, column=1, padx=5, pady=5)

        # File type selection
        self.lbl_file_type = tk.Label(self.root, text="Export file type:")
        self.lbl_file_type.grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.file_type_var = tk.StringVar(value=".txt")
        self.file_type_menu = ttk.Combobox(
            self.root,
            textvariable=self.file_type_var,
            values=[".txt", ".xls"],
            state="readonly",
        )
        self.file_type_menu.grid(row=5, column=1, padx=5, pady=5)


        self.btn_display = tk.Button(self.root, text="Display", command=self.display_data)
        self.btn_display.grid(row=6, column=0, padx=5, pady=5)
        self.btn_export = tk.Button(self.root, text="Export", command=self.export_data)
        self.btn_export.grid(row=6, column=1, padx=5, pady=5, sticky="w")
        self.btn_clear = tk.Button(self.root, text="Clear", command=self.clear_fields)
        self.btn_clear.grid(row=6, column=1, padx=5, pady=5, sticky="e")

    def load_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return

        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.sheet = self.workbook.active


            print("\nExcel content:")
            for row in self.sheet.iter_rows(values_only=True):
                print(row)


            first_row = next(self.sheet.iter_rows(values_only=True))
            print("\nfirst line:", first_row)

            if not first_row:
                messagebox.showerror("Error", "Excel file is empty!")
                return

            messagebox.showinfo("Success", "File uploaded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while uploading the file: {str(e)}\nDosya yolu: {file_path}")

    def display_data(self):
        if not hasattr(self, "sheet"):
            messagebox.showwarning("Warning", "Please upload a file first!")
            return

        student_id = self.entry_id.get()
        if not student_id.isdigit():
            messagebox.showerror("Error", "Enter a valid student ID.")
            return

        try:
            student_data = None
            print("\nSearched ID:", student_id)

            for row in self.sheet.iter_rows(values_only=True):
                print("Checked line:", row)
                if str(row[2]) == student_id:
                    student_data = row
                    print("student found:", student_data)
                    break

            if not student_data:
                messagebox.showinfo("Not Found", "Student ID not found!")
                return

            name = f"{student_data[0]} {student_data[1]}"
            scores = student_data[3:]


            print("Grades:", scores)
            if not scores or not all(isinstance(x, (int, float)) for x in scores if x is not None):
                messagebox.showerror("Error", "Grade data is invalid!")
                return

            gpa = round(sum(scores) / len(scores), 2)


            print(f"GPA calculation: {sum(scores)} / {len(scores)} = {gpa}")

            all_rows = list(self.sheet.iter_rows(values_only=True))
            all_gpas = []

            for row in all_rows[1:]:
                row_scores = row[3:]
                if all(isinstance(x, (int, float)) for x in row_scores if x is not None):
                    row_gpa = round(sum(row_scores) / len(row_scores), 2)
                    all_gpas.append(row_gpa)

            print("All GPAs:", all_gpas)

            rank = sorted(all_gpas, reverse=True).index(gpa) + 1
            print(f"Arrangement: {rank}")

            self.entry_name.config(state="normal")
            self.entry_name.delete(0, tk.END)
            self.entry_name.insert(0, name)
            self.entry_name.config(state="readonly")

            self.entry_gpa.config(state="normal")
            self.entry_gpa.delete(0, tk.END)
            self.entry_gpa.insert(0, gpa)
            self.entry_gpa.config(state="readonly")

            self.entry_rank.config(state="normal")
            self.entry_rank.delete(0, tk.END)
            self.entry_rank.insert(0, rank)
            self.entry_rank.config(state="readonly")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while displaying data: {str(e)}")
            print("Error detail:", e)

    def export_data(self):
        if not self.entry_name.get():
            messagebox.showwarning("Warning", "Please display student data first!")
            return


        student_id = self.entry_id.get()
        student_name = self.entry_name.get().replace(" ", "_")
        file_type = self.file_type_var.get()
        filename = f"{student_id}_{student_name}{file_type}"
        file_path = os.path.join(self.export_dir, filename)

        try:
            if file_type == ".txt":
                with open(file_path, "w") as file:
                    file.write(f"Student Report\n")
                    file.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    file.write(f"{'=' * 40}\n\n")
                    file.write(f"ID: {self.entry_id.get()}\n")
                    file.write(f"Name: {self.entry_name.get()}\n")
                    file.write(f"GPA: {self.entry_gpa.get()}\n")
                    file.write(f"Rank: {self.entry_rank.get()}\n")
            elif file_type == ".xls":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Student Report"

                headers = ["Field", "Value"]
                ws.append(headers)

                data = [
                    ["ID", self.entry_id.get()],
                    ["Name", self.entry_name.get()],
                    ["GPA", self.entry_gpa.get()],
                    ["Rank", self.entry_rank.get()],
                    ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                ]

                for row in data:
                    ws.append(row)

                wb.save(file_path)

            messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")

    def clear_fields(self):
        self.entry_id.delete(0, tk.END)

        self.entry_name.config(state="normal")
        self.entry_name.delete(0, tk.END)
        self.entry_name.config(state="readonly")

        self.entry_gpa.config(state="normal")
        self.entry_gpa.delete(0, tk.END)
        self.entry_gpa.config(state="readonly")

        self.entry_rank.config(state="normal")
        self.entry_rank.delete(0, tk.END)
        self.entry_rank.config(state="readonly")


if __name__ == "__main__":
    root = tk.Tk()
    app = StudentGPAApp(root)
    root.mainloop()